#! python

from openpyxl import Workbook

import datetime
import logging
import getDraftResults
import ffyahoo
import sys
sys.stdout.reconfigure(encoding='utf-8')

logging.basicConfig(level=logging.DEBUG)


# Add Player (via add, waiver, or trade) Transaction
def addPlayer(transaction, players):
    owner_name = transaction['owner']
    player_name = transaction['player']

    # If player does not exist, create new player
    if player_name not in players:
        players[player_name] = {'owner': '', 'drop_date': None, 'cost': 6}

    # Update player's owner and cost
    player = players[player_name]
    player['owner'] = owner_name
    if transaction['type'] == 'waiver' and player['drop_date'] is not None:
        # If a player was dropped more than 3 days ago, they are not eligible to stay the same cost
        if transaction['date'] - datetime.timedelta(days=3) > player['drop_date']:
            transaction['type'] = 'add'
    if transaction['type'] == 'add':
        if player['cost'] > 6:
            player['cost'] = 6


# Drop Player Transaction
def dropPlayer(transaction, players):
    player_name = transaction['player']

    # Update player's owner and drop date
    player = players[player_name]
    player['owner'] = ''
    player['drop_date'] = transaction['date']


def getKeepers(wb, year):
    players = {}
    rosters = {}
    owners = ffyahoo.getOwners(year)
    draft_results = getDraftResults.getDraftResults(year)
    transactions = ffyahoo.getTransactions(year)

    # Input drafted players into player list
    for player_cost, draft_round in enumerate(draft_results):
        for draft_player in draft_round:
            player_name = draft_player[0]
            owner_name = draft_player[1]
            for owner in owners:
                if owner_name in owner:
                    owner_name = owner
                    break
            players[player_name] = {'owner': owner_name, 'drop_date': None, 'cost': player_cost}

    # Process Transactions
    for transaction in reversed(transactions):
        if transaction['type'] == 'drop':
            dropPlayer(transaction, players)
        else:
            addPlayer(transaction, players)

    # Add players to rosters
    for curr_player in players:
        player = players[curr_player]
        if player['owner'] == '':
            continue
        if player['owner'] not in rosters:
            rosters[player['owner']] = [(curr_player, player['cost'])]
        else:
            rosters[player['owner']].append((curr_player, player['cost']))

    # Print rosters
    # for roster in rosters:
    #     curr_roster = rosters[roster]
    #     curr_roster = sorted(curr_roster, key=lambda k: k[1])
    #     print(roster)
    #     for player in curr_roster:
    #         print("Player: {}, Cost: {}".format(player[0], player[1]))

    ws = wb.create_sheet(year)
    ws.title = year
    ws.cell(row=1, column=1, value='Team')
    ws.cell(row=1, column=2, value='Player')
    ws.cell(row=1, column=3, value='Cost')
    curr_row = 2
    for roster in rosters:
        curr_col = 1
        curr_roster = rosters[roster]
        curr_roster = sorted(curr_roster, key=lambda k: k[1])
        ws.cell(row=curr_row, column=curr_col, value=roster)
        for player in curr_roster:
            curr_col = 2
            ws.cell(row=curr_row, column=curr_col, value=player[0])
            curr_col += 1
            ws.cell(row=curr_row, column=curr_col, value=player[1])
            curr_row += 1


def main():
    wb = Workbook()
    for year in range(2021, 2022):
        getKeepers(wb, str(year))
    wb.save('keepers.xlsx')


if __name__ == '__main__':
    main()
